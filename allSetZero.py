# -*- coding: utf-8 -*-
"""
Created on Tue Aug 22 18:10:47 2023

@author: eiru_
"""

from openpyxl import load_workbook

def zero_out_numbers(file_path, sheet_name):
    # Excelファイルを読み込みます
    wb = load_workbook(filename=file_path)
    ws = wb[sheet_name]

    # セルの値が数値の場合、0に変更します
    for i, row in enumerate(ws.iter_cols(values_only=False)):
        if i == 0 or i == 10:
            continue
        for cell_value in row:
            if isinstance(cell_value.value, (int, float)):
                # 変更後の値をセルに書き込む
                cell = ws.cell(row=cell_value.row, column=cell_value.column, value=0)

    # 変更内容を保存します
    wb.save(filename=file_path)

def count_of_numbers(file_path, sheet_name):
    # Excelファイルを読み込みます
    wb = load_workbook(filename=file_path)
    ws = wb[sheet_name]
    
    count_array = {}
    #print(count_array)
    # セルの値が数値の場合、0に変更します
    for i, row in enumerate(ws.iter_cols(values_only=False)):
        if i == 0 or i == 10:
            continue
        for j, cell_value in enumerate(row):
            if j < 2 or j > 50:
                continue

            if isinstance(cell_value.value, (int, float)):
                # 変更後の値をセルに書き込む
                count_array.setdefault(i,{})[j-1] = cell_value.value
   # print(count_array)
    # 変更内容を保存します
    return count_array