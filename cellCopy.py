# -*- coding: utf-8 -*-
"""
Created on Tue Aug 22 18:23:14 2023

@author: eiru_
"""

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Font, Alignment

def copy_cell_style(source_cell, destination_cell):
    # フォントのコピー
    if source_cell.font is not None:
        font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            color=source_cell.font.color
        )
        destination_cell.font = font

    # テキストの揃えのコピー
    if source_cell.alignment is not None:
        alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            wrap_text=source_cell.alignment.wrap_text
        )
        destination_cell.alignment = alignment

def copy_border_and_style(source_file, sheet_name, destination_file):
    # ソースファイルを読み込みます
    source_wb = load_workbook(filename=source_file)
    source_wb.active = source_wb[sheet_name]
    source_ws = source_wb.active

    # 新しいワークブックを作成します
    destination_wb = Workbook()
    #destination_wb.active =destination_wb.worksheets[1]
    destination_ws = destination_wb.active

    # セルの罫線とスタイルをコピーします
    for row in source_ws.iter_rows():
        for cell in row:
            destination_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.border is not None:
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                destination_ws.cell(row=cell.row, column=cell.column).border = border

            # セルのフォントと揃えをコピーします
            copy_cell_style(cell, destination_ws.cell(row=cell.row, column=cell.column))

    # 結果を保存します
    destination_wb.save(filename=destination_file)

if __name__ == "__main__":
    # ソースファイルと出力ファイルのパスを指定します
    source_file_path = "C:/Users/eiru_/Downloads/整備/Book1.xlsx"
    destination_file_path = "C:/Users/eiru_/Downloads/整備/test3.xlsx"
    sheet_name = "Sheet1"
    # 罫線とスタイルを含めてコピーした新しいExcelファイルを作成します
    copy_border_and_style(source_file_path, sheet_name, destination_file_path)
