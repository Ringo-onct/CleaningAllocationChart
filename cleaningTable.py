from openpyxl import load_workbook
import allSetZero
import random
import cellCopy
import datetime

def  gain_job(file_path, sheet_name):
    wb = load_workbook(filename=file_path)
    ws = wb[sheet_name]

    job_array = []
    job_dict = {}
    for i, row in enumerate(ws.iter_cols(values_only=False)):
        if i > 9:
            continue
        for j, cell_value in enumerate(row):
            if j < 2 or j > 49:
                continue

            if i == 0:
                job_array.append(cell_value.value)
                # 変更後の値をセルに書き込む
            else:
                cell_value_V = cell_value.value
                if cell_value_V == None:
                    cell_value_V = 0
                if job_array[j-2] == None:
                    break
                job_dict.setdefault(i,{})[job_array[j-2]] = cell_value_V
    return job_dict




def set_job(job_dict, chain_count):
    job_array = []
    set_dict = {}
    for floor,job in job_dict.items():
        job_array = []
        job_counter = 0
        for job_name,job_count in job.items():
            for _ in range(job_count):
                job_array.append(job_name)
                job_counter += 1
        job_array = random.sample(job_array, len(job_array))
        e = sorted(chain_count[floor].items(), key = lambda fruit : fruit[1])
        #print(len(e))
        #print(job_counter)
        e = shafle(e,e[job_counter-1][1])

        for i, j in enumerate(e):
            if i < len(job_array):
                set_dict.setdefault(floor,{})[j[0]] = job_array[i]
            else:
                set_dict.setdefault(floor,{})[j[0]] = "自室清掃"

   # print(set_dict)
    return set_dict

def write_job(file_path, set_dict):
    wb = load_workbook(filename=file_path)
    ws = wb.active
    for floor, room_job in set_dict.items():
        for room_namber, job in room_job.items():
            ws.cell(row=room_namber+2, column=floor+1, value=job)
    wb.save(filename=file_path)






def shafle(List,number):
    f = 0
    e = 0
    f_list = []
    e_list = []
    for x in List:
        if x[1] == number:
            f_list = List[:f]
            break
        else:
            f += 1
    for x in reversed(List):
        if x[1] == number:
            e_list = List[len(List)-e:]
            break
        else:
            e += 1
    e = len(List)-e
    h_list = f_list+random.sample(List[f:e], len(List[f:e]))+e_list
    return h_list




def al(file_path, sheet_name_1, sheet_name_2):

    destination_file_path = "割り振り票/"+str(datetime.date.today())+".xlsx"
    # 数字のセルを0に変更します
    cellCopy.copy_border_and_style(file_path, sheet_name_2, destination_file_path)
    job_dict = gain_job(file_path, sheet_name_1)
    c = allSetZero.count_of_numbers(file_path, sheet_name_2)
    set_dict = set_job(job_dict, c)
    write_job(destination_file_path, set_dict)

    return set_dict
