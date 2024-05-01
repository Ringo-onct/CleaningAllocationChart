from openpyxl import load_workbook
import cleaningTable
import collectGarbageRotation

def chain(file_path, sheet_name, job_list):
    wb = load_workbook(filename=file_path)
    ws = wb[sheet_name]
    for floor, room_job in job_list.items():
        for room_namber, job in room_job.items():
            cell_v = ws.cell(row=room_namber+2, column=floor+1)
            if job == "自室清掃" and cell_v.value <= 0:
                ws.cell(row=room_namber+2, column=floor+1,value=cell_v.value-1)
            elif job == "自室清掃" and cell_v.value > 0:
                ws.cell(row=room_namber+2, column=floor+1,value=0)
            else:
                ws.cell(row=room_namber+2, column=floor+1,value=int(cell_v.value+1))
    wb.save(filename=file_path)
    #print(cell_v.value)


def sum_seisou(file_path, sheet_name, job_list):
    wb = load_workbook(filename=file_path)
    ws = wb[sheet_name]
    for floor, room_job in job_list.items():
        for room_namber, job in room_job.items():
            cell_v = ws.cell(row=room_namber+2, column=floor+1)
            if job != "自室清掃":

                ws.cell(row=room_namber+2, column=floor+1,value=cell_v.value+1)
    wb.save(filename=file_path)
    #print(cell_v.value)

file_path_1 = "base.xlsx"
sheet_name_1 = "Sheet1"  # 適切なシート名に変更してください
sheet_name_2 = "Sheet2"
sheet_name_3 = "Sheet3"

collectGarbageRotation.sift_vaul(file_path_1, "Sheet4", sheet_name_1)   #清掃割り振り表の集積をローテーションさせる
job_list = cleaningTable.al(file_path_1, sheet_name_1, sheet_name_2)
sum_seisou(file_path_1, sheet_name_3, job_list)
chain(file_path_1, sheet_name_2, job_list)
